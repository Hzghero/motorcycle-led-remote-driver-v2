# -*- coding: utf-8 -*-
"""生成 发射端报价表.xlsx 与 接收端报价表.xlsx（含公式）。需安装 openpyxl: pip install openpyxl"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def bom_sheet(ws, rows, title):
    ws.append(["序号", "型号", "规格说明", "数量", "单价", "小计"])
    n = len(rows)
    for i, row in enumerate(rows, start=2):
        ws.append([i - 1, row[0], row[1], row[2], row[3], None])
        ws.cell(row=i, column=6).value = f"=D{i}*E{i}"
    # 元器件小计
    r_sum = n + 2
    ws.append(["合计", "", "元器件小计", "", "", f"=SUM(F2:F{n+1})"])
    # SMT 贴片加工费（下方）
    r_total_pts = r_sum + 1
    ws.append(["贴片元件总数量", "", "", None, "", ""])
    ws.cell(row=r_total_pts, column=4).value = f"=SUM(D2:D{n+1})"
    r_smt = r_sum + 2
    ws.append(["SMT贴片加工费", "按板计", 1, None, None, None])
    ws.cell(row=r_smt, column=6).value = f"=D{r_smt}*E{r_smt}"
    r_grand = r_sum + 3
    ws.append(["合计", "", "总计", "", "", f"=F{r_sum}+F{r_smt}"])
    for c in range(1, 7):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

def main():
    base = os.path.dirname(os.path.abspath(__file__))
    # 确保写入到脚本所在目录
    os.chdir(base)
    tx_rows = [
        ("STM32C011F4P6TR", "TSSOP-20 主控MCU", 1, None),
        ("SI24R1", "2.4G射频芯片", 1, None),
        ("16MHz无源晶振", "SI24R1时钟", 1, None),
        ("CR2032电池座", "纽扣电池座", 1, None),
        ("LDO 3.0V", "可选 XC6206等", 1, None),
        ("轻触开关", "按键 左/中/右", 3, None),
        ("LED 红色", "对码指示灯 低电流", 1, None),
        ("电阻 1k", "LED限流 0603", 1, None),
        ("电阻 10k", "上拉 0603", 3, None),
        ("电容 100nF", "去耦 0603", 4, None),
        ("电容 10µF", "电源滤波 0603/0805", 1, None),
        ("电容 22pF", "晶振负载", 2, None),
        ("SWD排针 2x5", "调试下载", 1, None),
        ("PCB天线", "板载2.4G天线", 1, None),
    ]
    rx_rows = [
        ("STM32C011F4P6TR", "TSSOP-20 主控MCU", 1, None),
        ("SI24R1", "2.4G射频接收", 1, None),
        ("LM3409", "LED恒流驱动 每路1颗", 2, None),
        ("DC-DC或LDO 12V转3.3V", "如MP2315/AMS1117-3.3", 1, None),
        ("电感", "DC-DC用 如4.7µH", 1, None),
        ("16MHz无源晶振", "SI24R1时钟", 1, None),
        ("电容 22pF", "晶振负载", 2, None),
        ("电容 100nF", "MCU/射频去耦 0603", 6, None),
        ("电容 10µF", "电源输入滤波", 2, None),
        ("电容 47µF", "12V输入滤波", 1, None),
        ("电阻 10k", "上拉/分压 0603", 4, None),
        ("电阻", "LM3409外围设定", 4, None),
        ("电容", "LM3409外围", 4, None),
        ("电感", "LM3409续流 每路1个", 2, None),
        ("接线端子", "电瓶/ACC/黄/蓝/白/天线/LED输出", 1, None),
        ("PCB天线或外接天线", "2.4G天线", 1, None),
    ]
    wb_tx = Workbook()
    bom_sheet(wb_tx.active, tx_rows, "发射端")
    wb_tx.save(os.path.join(base, "发射端报价表.xlsx"))
    wb_rx = Workbook()
    bom_sheet(wb_rx.active, rx_rows, "接收端")
    wb_rx.save(os.path.join(base, "接收端报价表.xlsx"))
    print("已生成: 发射端报价表.xlsx, 接收端报价表.xlsx")

if __name__ == "__main__":
    main()
